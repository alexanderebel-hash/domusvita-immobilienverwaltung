#!/usr/bin/env python3
"""
Import Bewohner from Excel (Airtable export) into PostgreSQL.

Usage:
    pip install openpyxl asyncpg python-dotenv
    python scripts/import_bewohner.py --file bewohner.xlsx [--dry-run]

Expected Excel columns:
    Name, Vorname, Wohngemeinschaft, Zimmer, Fester Einzugstermin,
    Status Klient, Kosten Pauschale, aktuell in Bearbeitung von

WG name mapping:
    "Sterndamm 10"        -> wg-sterndamm
    "Kupferkessel"         -> wg-kupferkessel
    "Kupferkessel Klein"   -> wg-kupferkessel-klein
    "Kupferkesselchen"     -> wg-kupferkessel-klein
    "Drachenwiese"         -> wg-drachenwiese
    "Drachenblick"         -> wg-drachenblick
"""

import argparse
import asyncio
import json
import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

# Add backend to path for database module
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

import asyncpg
from dotenv import load_dotenv

# Load .env from backend
load_dotenv(Path(__file__).parent.parent / "backend" / ".env")

# WG name mapping (case-insensitive)
WG_MAPPING = {
    "sterndamm": "wg-sterndamm",
    "sterndamm 10": "wg-sterndamm",
    "kupferkessel": "wg-kupferkessel",
    "kupferkessel groß": "wg-kupferkessel",
    "kupferkessel gross": "wg-kupferkessel",
    "kupferkessel klein": "wg-kupferkessel-klein",
    "kupferkesselchen": "wg-kupferkessel-klein",
    "drachenwiese": "wg-drachenwiese",
    "drachenblick": "wg-drachenblick",
}

# Status mapping from Excel -> DB
STATUS_MAPPING = {
    "01. interessent": "neu",
    "02. erstgespräch": "erstgespraech",
    "03. besichtigung": "besichtigung_geplant",
    "04. unterlagen": "unterlagen_gesendet",
    "05. entscheidung": "entscheidung_ausstehend",
    "06. zusage": "zusage",
    "07. klient ist eingezogen": "bewohner",
    "07. eingezogen": "bewohner",
    "08. auszug": "auszug_geplant",
    "09. ausgezogen": "ausgezogen",
    "10. verstorben": "verstorben",
    "abgesagt": "abgesagt",
}


def generate_id():
    return str(uuid.uuid4())


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def map_wg_name(name: str) -> str | None:
    if not name:
        return None
    normalized = name.strip().lower()
    for key, wg_id in WG_MAPPING.items():
        if key in normalized:
            return wg_id
    return None


def map_status(status_text: str) -> str:
    if not status_text:
        return "neu"
    normalized = status_text.strip().lower()
    for key, value in STATUS_MAPPING.items():
        if key in normalized:
            return value
    # Default: if "eingezogen" anywhere, it's bewohner
    if "eingezogen" in normalized:
        return "bewohner"
    return "neu"


def parse_date(val) -> str | None:
    if not val:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, str):
        # Try common formats
        for fmt in ["%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                return datetime.strptime(val.strip(), fmt).isoformat()
            except ValueError:
                continue
    return None


def parse_zimmer_nummer(val: str) -> str | None:
    if not val:
        return None
    val = str(val).strip()
    # Extract number from "Zimmer 3" or just "3"
    import re
    match = re.search(r'\d+', val)
    return match.group(0) if match else val


async def main(excel_path: str, dry_run: bool = False):
    print(f"{'[DRY RUN] ' if dry_run else ''}Import Bewohner from: {excel_path}")

    # Read Excel
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    # Find header row
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Found columns: {headers}")

    # Map column indices
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if "nachname" in h_lower or h_lower == "name":
            col_map["nachname"] = i
        elif "vorname" in h_lower:
            col_map["vorname"] = i
        elif "wohngemeinschaft" in h_lower or "wg" in h_lower:
            col_map["wg"] = i
        elif "zimmer" in h_lower:
            col_map["zimmer"] = i
        elif "einzug" in h_lower or "fester" in h_lower:
            col_map["einzugsdatum"] = i
        elif "status" in h_lower:
            col_map["status"] = i

    print(f"Column mapping: {col_map}")

    if "nachname" not in col_map or "vorname" not in col_map:
        print("ERROR: Could not find Name/Vorname columns")
        sys.exit(1)

    # Parse rows
    bewohner = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nachname = str(row[col_map["nachname"]] or "").strip()
        vorname = str(row[col_map.get("vorname", 0)] or "").strip()

        if not nachname or not vorname:
            continue  # Skip empty rows

        wg_name = str(row[col_map.get("wg", 0)] or "").strip() if "wg" in col_map else ""
        wg_id = map_wg_name(wg_name)
        zimmer_nr = parse_zimmer_nummer(row[col_map.get("zimmer", 0)]) if "zimmer" in col_map else None
        einzugsdatum = parse_date(row[col_map.get("einzugsdatum", 0)]) if "einzugsdatum" in col_map else None
        status = map_status(str(row[col_map.get("status", 0)] or "")) if "status" in col_map else "bewohner"

        bewohner.append({
            "nachname": nachname,
            "vorname": vorname,
            "wg_id": wg_id,
            "wg_name": wg_name,
            "zimmer_nr": zimmer_nr,
            "einzugsdatum": einzugsdatum,
            "status": status,
        })

    wb.close()
    print(f"\nParsed {len(bewohner)} rows from Excel")

    if dry_run:
        for b in bewohner:
            print(f"  {b['vorname']} {b['nachname']} -> WG: {b['wg_id']} / Zimmer: {b['zimmer_nr']} / Status: {b['status']}")
        print("\n[DRY RUN] No database changes made.")
        return

    # Connect to database
    pool = await asyncpg.create_pool(
        host=os.environ.get("PGHOST", "localhost"),
        port=int(os.environ.get("PGPORT", 5432)),
        database=os.environ.get("PGDATABASE", "domusvita"),
        user=os.environ.get("PGUSER", ""),
        password=os.environ.get("PGPASSWORD", ""),
        ssl="require" if os.environ.get("PGHOST", "").endswith("azure.com") else None,
        min_size=1,
        max_size=3,
    )

    created = 0
    skipped = 0
    errors = 0

    for b in bewohner:
        try:
            # Check if already exists (nachname + vorname + wg)
            existing = await pool.fetchval(
                "SELECT data->>'id' FROM klienten WHERE data->>'nachname' = $1 AND data->>'vorname' = $2",
                b["nachname"], b["vorname"]
            )

            if existing:
                print(f"  SKIP (exists): {b['vorname']} {b['nachname']} (id: {existing})")
                skipped += 1
                continue

            klient_id = generate_id()
            klient = {
                "id": klient_id,
                "vorname": b["vorname"],
                "nachname": b["nachname"],
                "status": b["status"],
                "pflegegrad": "keiner",
                "einzugsdatum": b["einzugsdatum"],
                "anfrage_quelle": "excel_import",
                "dringlichkeit": "flexibel",
                "bevorzugte_wgs": [b["wg_id"]] if b["wg_id"] else [],
                "anfrage_am": now_iso(),
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }

            # Insert klient
            await pool.execute(
                "INSERT INTO klienten (data) VALUES ($1::jsonb)",
                json.dumps(klient)
            )

            # Try to assign to room
            if b["wg_id"] and b["zimmer_nr"]:
                zimmer_row = await pool.fetchrow(
                    "SELECT data->>'id' as id, data->>'status' as status FROM wg_zimmer "
                    "WHERE data->>'pflege_wg_id' = $1 AND data->>'nummer' = $2",
                    b["wg_id"], b["zimmer_nr"]
                )
                if zimmer_row:
                    zimmer_id = zimmer_row["id"]
                    await pool.execute(
                        "UPDATE wg_zimmer SET data = data || $1::jsonb WHERE data->>'id' = $2",
                        json.dumps({
                            "status": "belegt",
                            "aktueller_bewohner_id": klient_id,
                            "updated_at": now_iso(),
                        }),
                        zimmer_id
                    )
                    await pool.execute(
                        "UPDATE klienten SET data = data || $1::jsonb WHERE data->>'id' = $2",
                        json.dumps({
                            "zimmer_id": zimmer_id,
                            "status": "bewohner",
                        }),
                        klient_id
                    )
                    print(f"  OK: {b['vorname']} {b['nachname']} -> {b['wg_id']}/Zimmer {b['zimmer_nr']}")
                else:
                    print(f"  WARN: Room not found: {b['wg_id']}/Zimmer {b['zimmer_nr']} for {b['vorname']} {b['nachname']}")
            else:
                print(f"  OK: {b['vorname']} {b['nachname']} (no room assignment)")

            created += 1

        except Exception as e:
            print(f"  ERROR: {b['vorname']} {b['nachname']}: {e}")
            errors += 1

    await pool.close()

    print(f"\nDone! Created: {created}, Skipped: {skipped}, Errors: {errors}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Bewohner from Excel to PostgreSQL")
    parser.add_argument("--file", required=True, help="Path to Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Only parse, don't write to DB")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    asyncio.run(main(args.file, args.dry_run))
